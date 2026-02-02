import openpyxl
import os

file_path = os.path.join("data", "Import_Template.xlsx")
wb = openpyxl.load_workbook(file_path, data_only=True)

print("Sheets:", wb.sheetnames)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"\n--- {sheet} ---")
    rows = list(ws.iter_rows(max_row=3, values_only=True))
    for i, row in enumerate(rows):
        print(f"Row {i+1}: {row}")
