import openpyxl
import os

file_path = os.path.join("data", "Test_Import_Template.xlsx")
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
else:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("Sheets:", wb.sheetnames)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(max_row=2, values_only=True))
        print(f"Sheet: {sheet}, Row 1 (Headers): {rows[0] if rows else 'Empty'}")
