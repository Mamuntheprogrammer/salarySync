import openpyxl
from datetime import datetime, time, date, timedelta
from sqlalchemy.orm import Session
from database import get_db_session
from models import (
    Employee, Shift, Attendance, Company, BusinessArea, 
    Designation, DesignationSubcategory, WeeklyHoliday, 
    LeaveQuota, HolidayCalendar, ShortLeave, Base
)
import os

class LegacyImportService:
    
    TABLE_SCHEMAS = {
        "companies": {
            "headers": ['code', 'name'],
            "dummy": [1000, 'APL']
        },
        "business_areas": {
            "headers": ['company_code', 'code', 'name'],
            "dummy": [1000, 10, 'Dhaka']
        },
        "designations": {
            "headers": ['name'],
            "dummy": ['Manager']
        },
        "designation_subcategories": {
            "headers": ['designation_name', 'name'],
            "dummy": ['Manager', 'Level-1']
        },
        "shifts": {
            "headers": ['name', 'start_time', 'end_time', 'late_allowance_minutes'],
            "dummy": ['Morning-1', '08:00 AM', '05:00 PM', 10]
        },
        "employees": {
            "headers": ['attendance_code', 'full_name', 'joining_date', 'salary_base', 'company_code', 'business_area_code', 'designation_name', 'subcategory_name', 'shift_name', 'custom_shift_start', 'custom_shift_end', 'is_active'],
            "dummy": [500001, 'Mehedi', '2022-05-12', 10000, 1000, 10, 'Manager', 'Level-1', 'Morning-1', None, None, True]
        },
        "attendance": {
            "headers": ['attendance_code', 'date', 'clock_in', 'clock_out'],
            "dummy": [500001, '2026-01-01', '10:21 AM', '08:45 PM']
        },
        "weekly_holidays": {
            "headers": ['day_of_week', 'company_code', 'business_area_code', 'shift_name'],
            "dummy": ['Friday', 1000, 10, 'Morning-1']
        },
        "leave_quotas": {
            "headers": ['year', 'leave_type', 'quota_limit', 'company_code', 'business_area_code'],
            "dummy": [2026, 'Annual', 14, 1000, 10]
        },
        "holiday_calendar": {
            "headers": ['date', 'description', 'is_ot_eligible', 'year', 'type', 'company_code', 'business_area_code'],
            "dummy": ['2026-01-15', 'Election ', False, 2026, 'Govment', 1000, 10]
        },
        "short_leaves": {
            "headers": ['attendance_code', 'date', 'start_time', 'end_time', 'reason', 'status'],
            "dummy": [500001, '2026-01-03', '10:00 AM', '12:00 PM', 'Personal', 'Approved']
        },
    }

    @staticmethod
    def generate_template(table_types, output_path):
        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        for table_type in table_types:
            schema = LegacyImportService.TABLE_SCHEMAS.get(table_type)
            if not schema:
                continue
                
            ws = wb.create_sheet(title=table_type)
            # Headers
            ws.append(schema["headers"])
            # Dummy Data
            ws.append(schema["dummy"])
            
        if not wb.sheetnames:
             # Fallback if no valid tables
             wb.create_sheet("Empty")
        
        wb.save(output_path)
        return output_path

    @staticmethod
    def import_table_data(file_path, table_type):
        session = get_db_session()
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        if table_type not in wb.sheetnames:
             return 0, [f"Sheet '{table_type}' not found in workbook."]
             
        ws = wb[table_type]
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return 0, ["Empty sheet"]
            
        headers = [str(h).strip().lower() for h in rows[0] if h]
        data_rows = rows[1:]
        
        importer = getattr(LegacyImportService, f"_import_{table_type.lower()}", None)
        if not importer:
            return 0, [f"No importer defined for {table_type}"]
            
        count = 0
        errors = []
        
        for idx, row in enumerate(data_rows, start=2):
            try:
                row_dict = {}
                # Map row by index to header
                for i, h in enumerate(headers):
                    if i < len(row):
                        row_dict[h] = row[i]
                
                if importer(session, row_dict):
                    count += 1
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        try:
            session.commit()
        except Exception as e:
            session.rollback()
            return 0, [f"Commit Error: {str(e)}"]
            
        return count, errors

    # --- Helpers ---
    @staticmethod
    def _parse_time(val):
        if val is None: return None
        if isinstance(val, time): return val
        if isinstance(val, datetime): return val.time()
        if isinstance(val, str):
            val = val.strip()
            formats = ["%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"]
            for f in formats:
                try: return datetime.strptime(val, f).time()
                except: continue
        return None

    @staticmethod
    def _parse_date(val):
        if val is None: return None
        if isinstance(val, date): return val
        if isinstance(val, datetime): return val.date()
        if isinstance(val, str):
            val = val.strip()
            formats = ["%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"]
            for f in formats:
                try: return datetime.strptime(val, f).date()
                except: continue
        return None

    @staticmethod
    def _get_bool(val):
        if isinstance(val, bool): return val
        if isinstance(val, str):
            return val.lower() in ['true', '1', 'yes', 'y']
        return bool(val)

    # --- Importers ---

    @staticmethod
    def _import_companies(session, data):
        code = str(data.get('code')).strip()
        if not code or code.lower() == 'none': return False
        
        name = data.get('name')
        if not name: raise ValueError("Missing name")
        
        company = session.query(Company).filter_by(code=code).first()
        if not company:
            company = Company(code=code, name=name)
            session.add(company)
        else:
            company.name = name # update
        return True

    @staticmethod
    def _import_business_areas(session, data):
        comp_code = str(data.get('company_code')).strip()
        code = str(data.get('code')).strip()
        if not code or code.lower() == 'none': return False
        
        name = data.get('name')
        
        company = session.query(Company).filter_by(code=comp_code).first()
        if not company: raise ValueError(f"Company {comp_code} not found")
        
        ba = session.query(BusinessArea).filter_by(code=code, company_id=company.id).first()
        if not ba:
            ba = BusinessArea(code=code, name=name, company_id=company.id)
            session.add(ba)
        else:
            ba.name = name
        return True

    @staticmethod
    def _import_designations(session, data):
        name = data.get('name')
        if not name or str(name).lower() == 'none': return False
        
        desig = session.query(Designation).filter_by(name=name).first()
        if not desig:
            desig = Designation(name=name)
            session.add(desig)
        return True

    @staticmethod
    def _import_designation_subcategories(session, data):
        des_name = data.get('designation_name')
        if not des_name or str(des_name).lower() == 'none': return False
        
        sub_name = data.get('name')
        
        desig = session.query(Designation).filter_by(name=des_name).first()
        # If designation not found, maybe create it? Or strict? Let's be strict.
        if not desig: raise ValueError(f"Designation {des_name} not found")
        
        sub = session.query(DesignationSubcategory).filter_by(name=sub_name, designation_id=desig.id).first()
        if not sub:
            sub = DesignationSubcategory(name=sub_name, designation_id=desig.id)
            session.add(sub)
        return True

    @staticmethod
    def _import_shifts(session, data):
        name = data.get('name')
        if not name or str(name).lower() == 'none': return False
        
        start = LegacyImportService._parse_time(data.get('start_time'))
        end = LegacyImportService._parse_time(data.get('end_time'))
        late = int(data.get('late_allowance_minutes') or 15)
        
        shift = session.query(Shift).filter_by(name=name).first()
        if not shift:
            shift = Shift(name=name, start_time=start, end_time=end, late_allowance_minutes=late)
            session.add(shift)
        else:
            # Update
            shift.start_time = start
            shift.end_time = end
            shift.late_allowance_minutes = late
        return True

    @staticmethod
    def _import_employees(session, data):
        code_raw = data.get('attendance_code')
        if code_raw is None: 
            return False # Skip empty rows
            
        code = str(code_raw).strip()
        if not code or code.lower() == 'none': 
            raise ValueError("Invalid attendance code")

        full_name = data.get('full_name')
        
        emp = session.query(Employee).filter_by(attendance_code=code).first()
        if not emp:
            if not full_name:
                raise ValueError(f"Missing full_name for new employee code {code}")
            emp = Employee(attendance_code=code, full_name=full_name)
            session.add(emp)
            
        emp.full_name = data.get('full_name', emp.full_name)
        emp.joining_date = LegacyImportService._parse_date(data.get('joining_date')) or emp.joining_date
        emp.salary_base = float(data.get('salary_base') or 0)
        emp.is_active = LegacyImportService._get_bool(data.get('is_active', True))
        
        # Relations
        comp_code = str(data.get('company_code', '')).strip()
        if comp_code:
            c = session.query(Company).filter_by(code=comp_code).first()
            if c: emp.company_id = c.id
            
        ba_code = str(data.get('business_area_code', '')).strip()
        if ba_code:
            # Prefer finding by company? BA codes might duplicate across companies?
            # Model definition says BA code is NOT unique globally, but unique within company? 
            # Actually model just says Code String(2). It identifies by company_id.
            # So we need company to find BA safely.
            if emp.company_id:
                b = session.query(BusinessArea).filter_by(code=ba_code, company_id=emp.company_id).first()
                if b: emp.business_area_id = b.id
                
        des_name = data.get('designation_name')
        if des_name:
            d = session.query(Designation).filter_by(name=des_name).first()
            if d: emp.designation_id = d.id
            
        sub_name = data.get('subcategory_name')
        if sub_name and emp.designation_id:
            s = session.query(DesignationSubcategory).filter_by(name=sub_name, designation_id=emp.designation_id).first()
            if s: emp.designation_subcategory_id = s.id
            
        shift_name = data.get('shift_name')
        if shift_name:
            sh = session.query(Shift).filter_by(name=shift_name).first()
            if sh: emp.shift_id = sh.id
            
        # Custom Shift
        cs_start = LegacyImportService._parse_time(data.get('custom_shift_start'))
        cs_end = LegacyImportService._parse_time(data.get('custom_shift_end'))
        if cs_start: emp.custom_shift_start = cs_start
        if cs_end: emp.custom_shift_end = cs_end
        
        return True

    @staticmethod
    def _import_attendance(session, data):
        code_raw = data.get('attendance_code')
        if not code_raw or str(code_raw).lower() == 'none': return False
            
        code = str(code_raw).strip()
        date_val = LegacyImportService._parse_date(data.get('date'))
        
        if not date_val: raise ValueError("Missing date")
        
        emp = session.query(Employee).filter_by(attendance_code=code).first()
        if not emp: raise ValueError(f"Employee {code} not found")
        
        att = session.query(Attendance).filter_by(employee_id=emp.id, date=date_val).first()
        if not att:
            att = Attendance(employee_id=emp.id, date=date_val)
            session.add(att)
            
        cin = LegacyImportService._parse_time(data.get('clock_in'))
        cout = LegacyImportService._parse_time(data.get('clock_out'))
        
        if cin: att.clock_in = datetime.combine(date_val, cin)
        if cout: 
            # Handle Next Day crossover? If cout < cin?
            dt_out = datetime.combine(date_val, cout)
            if cin and cout < cin:
                dt_out += timedelta(days=1)
            att.clock_out = dt_out
            
        return True

    @staticmethod
    def _import_weekly_holidays(session, data):
        day_raw = data.get('day_of_week')
        if not day_raw or str(day_raw).lower() == 'none': return False

        company_code = str(data.get('company_code', '')).strip()
        ba_code = str(data.get('business_area_code', '')).strip()
        shift_name = data.get('shift_name')
        day_str = str(day_raw).strip().title()
        
        days_map = {"Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3, "Friday": 4, "Saturday": 5, "Sunday": 6}
        day_num = days_map.get(day_str)
        if day_num is None:
            if day_str.isdigit() and 0 <= int(day_str) <= 6:
                day_num = int(day_str)
            else:
                return False # specific logic: invalid day -> skip? or error? Let's skip if strictly invalid/empty row. But user said invalid day error. If row is empty, day_str is "None". Checked above. if it's "InvalidDay" raise error.
                raise ValueError(f"Invalid day: {day_str}")
        
        # Verify entities
        comp_id = None
        ba_id = None
        shift_id = None
        
        if company_code:
            c = session.query(Company).filter_by(code=company_code).first()
            if c: comp_id = c.id
        
        if ba_code and comp_id:
             b = session.query(BusinessArea).filter_by(code=ba_code, company_id=comp_id).first()
             if b: ba_id = b.id
             
        if shift_name:
            s = session.query(Shift).filter_by(name=shift_name).first()
            if s: shift_id = s.id
            
        # Check existing to avoid dupes
        q = session.query(WeeklyHoliday).filter_by(day_of_week=day_num)
        if comp_id: q = q.filter_by(company_id=comp_id)
        else: q = q.filter(WeeklyHoliday.company_id == None)
        
        if ba_id: q = q.filter_by(business_area_id=ba_id)
        else: q = q.filter(WeeklyHoliday.business_area_id == None)
        
        if shift_id: q = q.filter_by(shift_id=shift_id)
        else: q = q.filter(WeeklyHoliday.shift_id == None)
        
        if not q.first():
            wh = WeeklyHoliday(day_of_week=day_num, company_id=comp_id, business_area_id=ba_id, shift_id=shift_id)
            session.add(wh)
            
        return True

    @staticmethod
    def _import_leave_quotas(session, data):
        year_val = data.get('year')
        if not year_val or str(year_val).lower() == 'none': return False

        year = int(year_val)
        l_type = data.get('leave_type')
        limit = float(data.get('quota_limit') or 0)
        company_code = str(data.get('company_code', '')).strip()
        ba_code = str(data.get('business_area_code', '')).strip()
        
        comp_id = None
        ba_id = None
        
        if company_code:
            c = session.query(Company).filter_by(code=company_code).first()
            if c: comp_id = c.id
            
        if ba_code and comp_id:
             b = session.query(BusinessArea).filter_by(code=ba_code, company_id=comp_id).first()
             if b: ba_id = b.id
             
        # Check existing
        q = session.query(LeaveQuota).filter_by(year=year, leave_type=l_type)
        if comp_id: q = q.filter_by(company_id=comp_id)
        else: q = q.filter(LeaveQuota.company_id == None)
        
        if ba_id: q = q.filter_by(business_area_id=ba_id)
        else: q = q.filter(LeaveQuota.business_area_id == None)
        
        lq = q.first()
        if not lq:
            lq = LeaveQuota(year=year, leave_type=l_type, quota_limit=limit, company_id=comp_id, business_area_id=ba_id)
            session.add(lq)
        else:
            lq.quota_limit = limit
            
        return True

    @staticmethod
    def _import_holiday_calendar(session, data):
        date_raw = data.get('date')
        if not date_raw or str(date_raw).lower() == 'none': return False

        date_val = LegacyImportService._parse_date(date_raw)
        desc = data.get('description')
        if not date_val or not desc: raise ValueError("Missing date or description")
        
        is_ot = LegacyImportService._get_bool(data.get('is_ot_eligible'))
        year = int(data.get('year') or date_val.year)
        h_type = data.get('type')
        
        company_code = str(data.get('company_code', '')).strip()
        ba_code = str(data.get('business_area_code', '')).strip()
        
        comp_id = None
        ba_id = None
        
        if company_code and company_code.lower() != 'none':
            c = session.query(Company).filter_by(code=company_code).first()
            if c: 
                comp_id = c.id
            else:
                raise ValueError(f"Company {company_code} not found")
            
        if ba_code and ba_code.lower() != 'none':
            if not comp_id:
                 raise ValueError("Business Area requires Company Code to be specified")
            b = session.query(BusinessArea).filter_by(code=ba_code, company_id=comp_id).first()
            if b: 
                ba_id = b.id
            else:
                 raise ValueError(f"Business Area {ba_code} not found for Company {company_code}")
        
        h = session.query(HolidayCalendar).filter_by(
            date=date_val, 
            company_id=comp_id, 
            business_area_id=ba_id
        ).first()
        
        if not h:
            h = HolidayCalendar(
                date=date_val,
                description=desc,
                is_ot_eligible=is_ot,
                year=year,
                type=h_type,
                company_id=comp_id,
                business_area_id=ba_id
            )
            session.add(h)
        else:
            h.description = desc
            h.is_ot_eligible = is_ot
            h.type = h_type
            
        return True

    @staticmethod
    def _import_short_leaves(session, data):
        code_raw = data.get('attendance_code')
        if not code_raw or str(code_raw).lower() == 'none': return False
        
        code = str(code_raw).strip()
        date_val = LegacyImportService._parse_date(data.get('date'))
        start = LegacyImportService._parse_time(data.get('start_time'))
        end = LegacyImportService._parse_time(data.get('end_time'))
        
        if not date_val or not start or not end: 
            raise ValueError("Missing date, start_time or end_time")
            
        emp = session.query(Employee).filter_by(attendance_code=code).first()
        if not emp: raise ValueError(f"Employee {code} not found")
        
        reason = data.get('reason', '')
        status = data.get('status', 'Pending').capitalize()
        if status not in ['Pending', 'Approved', 'Rejected']: status = 'Pending'
        
        sl = session.query(ShortLeave).filter_by(employee_id=emp.id, date=date_val, start_time=start).first()
        if not sl:
            sl = ShortLeave(
                employee_id=emp.id,
                date=date_val,
                start_time=start,
                end_time=end,
                reason=reason,
                status=status
            )
            session.add(sl)
        else:
            sl.end_time = end
            sl.reason = reason
            sl.status = status
            
        return True
