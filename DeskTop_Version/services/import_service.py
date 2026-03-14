import openpyxl
from openpyxl.styles import Font
from sqlalchemy.orm import Session
from database import get_db_session
from models import Company, BusinessArea, Shift, Designation, DesignationSubcategory, Employee, HolidayCalendar, LeaveQuota, Attendance, ShortLeave, SalaryBreakdown
import os

from datetime import datetime, time, date

class ImportService:
    @staticmethod
    def generate_template(file_path):
        """Generates an Excel template with 10 rows of static sample data."""
        wb = openpyxl.Workbook()
        
        # Helper
        def create_sheet(name, headers, samples):
            if name in wb.sheetnames:
                ws = wb[name]
            else:
                ws = wb.create_sheet(name)
            
            if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb['Sheet']
                
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                
            for r_idx, sample_row in enumerate(samples, 2):
                for col_num, val in enumerate(sample_row, 1):
                    ws.cell(row=r_idx, column=col_num).value = val
                
        # 1. Companies
        create_sheet("Companies", ["code", "name"], [
            ["HQ01", "Global Tech Solutions"], ["HQ02", "Innovate Corp"], ["BR01", "Alpha Branch"], ["BR02", "Beta Branch"], ["SUB01", "Subsidiary A"],
            ["SUB02", "Subsidiary B"], ["JV01", "Joint Venture X"], ["JV02", "Joint Venture Y"], ["INT01", "International Div"], ["LOC01", "Local Operations"]
        ])
        
        # 2. BusinessAreas
        create_sheet("BusinessAreas", ["company_code", "area_code", "name"], [
            ["HQ01", "HR", "Human Resources"], ["HQ01", "IT", "Information Technology"], ["HQ01", "OPS", "Operations"], ["HQ01", "FIN", "Finance"], ["HQ02", "MKT", "Marketing"],
            ["HQ02", "SAL", "Sales"], ["HQ02", "SUP", "Support"], ["HQ02", "R&D", "Research"], ["BR01", "LOG", "Logistics"], ["BR01", "SEC", "Security"]
        ])
        
        # 3. Shifts
        create_sheet("Shifts", ["name", "start_time (HH:MM)", "end_time (HH:MM)", "late_allowance_min"], [
            ["General", "09:00", "18:00", 15], ["Morning", "06:00", "14:00", 10], ["Evening", "14:00", "22:00", 10], ["Night", "22:00", "06:00", 15], ["Early Bird", "05:00", "13:00", 5],
            ["Late Shift", "11:00", "20:00", 20], ["Weekend A", "10:00", "16:00", 0], ["Weekend B", "12:00", "18:00", 0], ["Split 1", "08:00", "17:00", 15], ["Split 2", "10:00", "19:00", 15]
        ])
        
        # 4. Designations
        create_sheet("Designations", ["name"], [
            ["Manager"], ["Developer"], ["QA Engineer"], ["Designer"], ["HR Executive"], ["Accountant"], ["Sales Rep"], ["Team Lead"], ["Director"], ["Intern"]
        ])
        
        # 5. Subcategories
        create_sheet("Subcategories", ["designation_name", "subcategory_name"], [
            ["Manager", "Senior"], ["Manager", "Junior"], ["Developer", "Frontend"], ["Developer", "Backend"], ["Developer", "Fullstack"],
            ["QA Engineer", "Automation"], ["QA Engineer", "Manual"], ["Sales Rep", "Field"], ["Sales Rep", "In-House"], ["Designer", "UI/UX"]
        ])
        
        # 6. Employees
        create_sheet("Employees", ["emp_id", "full_name", "company_code", "area_code", "shift_name", "designation_name", "subcategory_name", "salary_base", "is_active"], [
            ["1", "John Doe", "HQ01", "HR", "General", "Manager", "Senior", 50000, True],
            ["2", "Jane Smith", "HQ01", "IT", "Morning", "Developer", "Backend", 45000, True],
            ["3", "Alice Johnson", "HQ01", "IT", "Morning", "Developer", "Frontend", 42000, True],
            ["4", "Bob Brown", "HQ02", "MKT", "Evening", "Manager", "Junior", 35000, True],
            ["5", "Charlie Davis", "HQ02", "SAL", "Evening", "Sales Rep", "Field", 30000, True],
            ["6", "Eva Wilson", "HQ02", "SUP", "Night", "Team Lead", "", 40000, True],
            ["7", "Frank Miller", "BR01", "LOG", "General", "Intern", "", 15000, True],
            ["8", "Grace Lee", "BR01", "SEC", "Night", "Director", "Senior", 90000, True],
            ["9", "Henry Taylor", "HQ01", "FIN", "General", "Accountant", "", 38000, False],
            ["", "Ivy Clark", "HQ01", "IT", "Split 1", "Designer", "UI/UX", 41000, True]
        ])
        
        # 7. HolidayCalendar
        # Fields: name, date (YYYY-MM-DD), type, is_ot_eligible, year, company_code, business_area_code
        create_sheet("HolidayCalendar", ["name", "date (YYYY-MM-DD)", "type", "is_ot_eligible", "year", "company_code", "business_area_code"], [
            ["New Year", "2024-01-01", "National", True, 2024, "", ""],
            ["Victory Day", "2024-12-16", "National", True, 2024, "", ""],
            ["Labor Day", "2024-05-01", "National", True, 2024, "HQ01", ""],
            ["Eid Ul Fitr", "2024-04-10", "Festival", False, 2024, "", ""],
            ["Eid Ul Adha", "2024-06-17", "Festival", False, 2024, "", ""],
            ["Christmas", "2024-12-25", "Festival", True, 2024, "", ""],
            ["Company Founding Day (HQ01)", "2024-08-15", "Company", True, 2024, "HQ01", ""],
            ["Durga Puja", "2024-10-13", "Festival", True, 2024, "", ""],
            ["Independence Day", "2024-03-26", "National", True, 2024, "", ""],
            ["Language Day", "2024-02-21", "National", True, 2024, "", ""]
        ])
        
        # 8. LeaveQuota
        # Fields: emp_id, leave_type, limit (days/hours), year
        create_sheet("LeaveQuota", ["emp_id", "leave_type", "limit", "year"], [
            ["1", "Annual", 14, 2024], ["1", "Sick", 10, 2024], ["2", "Annual", 14, 2024], ["2", "Sick", 10, 2024],
            ["3", "Annual", 14, 2024], ["3", "Sick", 10, 2024], ["4", "Annual", 14, 2024], ["4", "Sick", 10, 2024],
            ["5", "Annual", 14, 2024], ["5", "Sick", 10, 2024]
        ])
        
        # 9. Attendance
        # Fields: emp_id, date (YYYY-MM-DD), clock_in (HH:MM), clock_out (HH:MM)
        create_sheet("Attendance", ["emp_id", "date (YYYY-MM-DD)", "clock_in (HH:MM)", "clock_out (HH:MM)"], [
            ["1", "2024-01-01", "09:05", "18:10"], ["1", "2024-01-02", "08:55", "18:05"], ["1", "2024-01-03", "09:10", "18:15"],
            ["2", "2024-01-01", "06:05", "14:05"], ["2", "2024-01-02", "06:00", "14:10"], ["2", "2024-01-03", "06:02", "14:00"],
            ["3", "2024-01-01", "09:00", "18:00"], ["3", "2024-01-02", "09:00", "18:00"], ["3", "2024-01-03", "09:00", "18:00"],
            ["4", "2024-01-01", "14:10", "22:15"]
        ])
        
        # 10. ShortLeave
        # Fields: emp_id, date (YYYY-MM-DD), start_time (HH:MM), end_time (HH:MM), reason
        create_sheet("ShortLeave", ["emp_id", "date (YYYY-MM-DD)", "start_time (HH:MM)", "end_time (HH:MM)", "reason"], [
            ["1", "2024-01-05", "10:00", "11:00", "Doctor appointment"],
            ["2", "2024-01-06", "14:00", "15:30", "Bank work"],
            ["3", "2024-01-07", "09:30", "10:00", "Personal errand"],
            ["4", "2024-01-08", "11:00", "12:00", "Emergency"],
            ["1", "2024-01-10", "15:00", "16:00", "Family matter"],
            ["2", "2024-01-11", "10:30", "11:30", "Medical checkup"],
            ["3", "2024-01-12", "13:00", "14:00", "Dental"],
            ["4", "2024-01-14", "09:00", "09:30", "Late arrival"],
            ["1", "2024-01-15", "16:00", "17:00", "Early exit"],
            ["5",  "2024-01-16", "10:00", "11:00", "Example"]
        ])
        
        # 11. SalaryBreakdown
        create_sheet("SalaryBreakdown", [
            "emp_id", "year", "valid_to (YYYY-MM-DD)", "basic", "house_rent_allowance", "conveyance", 
            "medical", "mobile_bill", "transportation_allowance", "other_allowance"
        ], [
            ["1", 2024, "2024-12-31", 20000, 10000, 5000, 3000, 1000, 2000, 9000],
            ["2", 2024, "", 18000, 9000, 4000, 2000, 500, 1000, 10500]
        ])
        
        wb.save(file_path)
        return True

    @staticmethod
    def get_sheet_names(file_path):
        """Returns list of sheet names from the excel file."""
        if not os.path.exists(file_path):
            return []
        
        wb = openpyxl.load_workbook(file_path, read_only=True)
        return wb.sheetnames

    @staticmethod
    def import_data(file_path, selected_sheets):
        """
        Imports data from selected sheets.
        Returns (success_count, errors_list)
        """
        session = get_db_session()
        wb = openpyxl.load_workbook(file_path, data_only=True)
        errors = []
        count = 0
        
        try:
            # 1. Companies
            if "Companies" in selected_sheets and "Companies" in wb.sheetnames:
                ws = wb["Companies"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]: continue # Skip empty
                    code, name = row[0], row[1]
                    
                    if not session.query(Company).filter_by(code=code).first():
                        c = Company(code=code, name=name)
                        session.add(c)
                        count += 1
                session.flush()

            # 2. BusinessAreas
            if "BusinessAreas" in selected_sheets and "BusinessAreas" in wb.sheetnames:
                ws = wb["BusinessAreas"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0] or not row[1]: continue
                    comp_code, area_code, name = row[0], row[1], row[2]
                    
                    comp = session.query(Company).filter_by(code=str(comp_code)).first()
                    if comp:
                        if not session.query(BusinessArea).filter_by(code=str(area_code), company_id=comp.id).first():
                            ba = BusinessArea(code=str(area_code), name=name, company_id=comp.id)
                            session.add(ba)
                            count += 1
                session.flush()
            
            # 3. Shifts
            if "Shifts" in selected_sheets and "Shifts" in wb.sheetnames:
                from datetime import datetime, time
                ws = wb["Shifts"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]: continue
                    name, start_str, end_str, late = row[0], row[1], row[2], row[3]
                    
                    # Basic parser for "HH:MM" or "HH:MM:SS"
                    def parse_time(t_val):
                        if isinstance(t_val, time): return t_val
                        if isinstance(t_val, datetime): return t_val.time()
                        if isinstance(t_val, str):
                            try: return datetime.strptime(t_val, "%H:%M").time()
                            except: 
                                try: return datetime.strptime(t_val, "%H:%M:%S").time()
                                except: return None
                        return None

                    s_time = parse_time(start_str)
                    e_time = parse_time(end_str)
                    
                    if s_time and e_time:
                         if not session.query(Shift).filter_by(name=name).first():
                             s = Shift(name=name, start_time=s_time, end_time=e_time, late_allowance_minutes=int(late or 15))
                             session.add(s)
                             count += 1
                session.flush()
                
            # 4. Designations
            if "Designations" in selected_sheets and "Designations" in wb.sheetnames:
                ws = wb["Designations"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]: continue
                    name = row[0]
                    if not session.query(Designation).filter_by(name=name).first():
                        session.add(Designation(name=name))
                        count += 1
                session.flush()
                
            # 5. Subcategories
            if "Subcategories" in selected_sheets and "Subcategories" in wb.sheetnames:
                ws = wb["Subcategories"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0] or not row[1]: continue
                    d_name, sub_name = row[0], row[1]
                    
                    deg = session.query(Designation).filter_by(name=d_name).first()
                    if deg:
                        if not session.query(DesignationSubcategory).filter_by(name=sub_name, designation_id=deg.id).first():
                            session.add(DesignationSubcategory(name=sub_name, designation_id=deg.id))
                            count += 1
                session.flush()

            # 6. Employees
            if "Employees" in selected_sheets and "Employees" in wb.sheetnames:
                ws = wb["Employees"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]: continue
                    # emp_id, full_name, company_code, area_code, shift_name, designation_name, subcategory_name, salary_base, is_active
                    emp_id_val = str(row[0]).strip() if row[0] else ""
                    fname = row[1]
                    ccode = str(row[2])
                    acode_ba = str(row[3])
                    sname = row[4]
                    dname = row[5]
                    subname = row[6]
                    salary = float(row[7] or 0)
                    active = bool(row[8]) if row[8] is not None else True
                    
                    # Resolve IDs
                    comp = session.query(Company).filter_by(code=ccode).first()
                    if not comp: 
                        errors.append(f"Employee {fname}: Company {ccode} not found")
                        continue
                        
                    ba = session.query(BusinessArea).filter_by(code=acode_ba, company_id=comp.id).first()
                    if not ba:
                        errors.append(f"Employee {fname}: Area {acode_ba} not found")
                        continue
                        
                    shift = session.query(Shift).filter_by(name=sname).first() if sname else None
                    deg = session.query(Designation).filter_by(name=dname).first() if dname else None
                    sub = None
                    if deg and subname:
                        sub = session.query(DesignationSubcategory).filter_by(name=subname, designation_id=deg.id).first()
                    
                    # Upsert or Insert check
                    emp = None
                    if emp_id_val and emp_id_val.isdigit():
                        emp = session.query(Employee).filter_by(id=int(emp_id_val)).first()
                        
                    if emp:
                        # Update?
                        emp.full_name = fname
                        emp.company_id = comp.id
                        emp.business_area_id = ba.id
                        emp.shift_id = shift.id if shift else None
                        emp.designation_id = deg.id if deg else None
                        emp.designation_subcategory_id = sub.id if sub else None
                        emp.salary_base = salary
                        emp.is_active = active
                        count += 1 # Count updates too?
                    else:
                        emp = Employee(
                            full_name=fname,
                            company_id=comp.id,
                            business_area_id=ba.id,
                            shift_id=shift.id if shift else None,
                            designation_id=deg.id if deg else None,
                            designation_subcategory_id=sub.id if sub else None,
                            salary_base=salary,
                            is_active=active
                        )
                        session.add(emp)
                        count += 1
                        
                        
            # 7. HolidayCalendar
            if "HolidayCalendar" in selected_sheets and "HolidayCalendar" in wb.sheetnames:
                ws = wb["HolidayCalendar"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]: continue
                    # name, date (YYYY-MM-DD), type, is_ot_eligible, year, company_code, business_area_code
                    name = row[0]
                    date_val = None
                    if isinstance(row[1], datetime): date_val = row[1].date()
                    elif isinstance(row[1], date): date_val = row[1]
                    elif isinstance(row[1], str):
                        try: date_val = datetime.strptime(row[1], "%Y-%m-%d").date()
                        except: pass
                    
                    h_type = row[2] or "National"
                    is_ot = bool(row[3]) if row[3] is not None else False
                    
                    comp_code = None
                    ba_code = None
                    # Try to read extra columns if they exist
                    if len(row) > 4: 
                        # Year is usually row[4] based on my template I generated earlier? 
                        # Wait, in generate_template I have: ["name", "date", "type", "is_ot_eligible"]
                        # The USER asked for: "Date,Description,Type,Is_OT_Eligible,Year,Company_Code,Business_Area_Code" in CSV.
                        # But here I am editing Excel Import.
                        # I must match the Excel columns I am about to define in generate_template.
                        # Let's assume I will update generate_template to have:
                        # name, date, type, is_ot_eligible, year, company_code, business_area_code
                        # So:
                        # 0: name
                        # 1: date
                        # 2: type
                        # 3: is_ot
                        # 4: year (optional?)
                        # 5: company_code
                        # 6: ba_code
                        pass
                    
                    # Safe reading
                    year_val = date_val.year if date_val else None
                    if len(row) > 4: year_val = row[4] or (date_val.year if date_val else None)
                    if len(row) > 5: comp_code = row[5]
                    if len(row) > 6: ba_code = row[6]

                    if date_val:
                         # Use filter with codes
                         # Note: Optional uniqueness check? 
                         # We should check if holiday exists for this specific scope.
                         
                         exists_query = session.query(HolidayCalendar).filter(
                             HolidayCalendar.date == date_val,
                             HolidayCalendar.company_code == comp_code,
                             HolidayCalendar.business_area_code == ba_code
                         )
                         
                         if not exists_query.first():
                             h = HolidayCalendar(
                                 description=name, 
                                 date=date_val, 
                                 type=h_type, 
                                 is_ot_eligible=is_ot, 
                                 year=int(year_val) if year_val else date_val.year,
                                 company_code=comp_code,
                                 business_area_code=ba_code
                             )
                             session.add(h)
                             count += 1
                session.flush()

            # 8. LeaveQuota
            if "LeaveQuota" in selected_sheets and "LeaveQuota" in wb.sheetnames:
                ws = wb["LeaveQuota"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]: continue
                    emp = None
                    if emp_id_val and emp_id_val.isdigit():
                        emp = session.query(Employee).filter_by(id=int(emp_id_val)).first()
                    if emp:
                         # Check if policy exists for this company/ba/year/type
                         q = session.query(LeaveQuota).filter_by(
                             company_id=emp.company_id, 
                             business_area_id=emp.business_area_id,
                             year=year,
                             leave_type=l_type
                         ).first()
                         if not q:
                             q = LeaveQuota(
                                 company_id=emp.company_id,
                                 business_area_id=emp.business_area_id,
                                 year=year,
                                 leave_type=l_type,
                                 quota_limit=limit
                             )
                             session.add(q)
                             count += 1
                session.flush()

            # 9. Attendance
            if "Attendance" in selected_sheets and "Attendance" in wb.sheetnames:
                ws = wb["Attendance"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0] and not row[1]: continue
                    # emp_id, date (YYYY-MM-DD), clock_in (HH:MM), clock_out (HH:MM)
                    emp_id_val = str(row[0]).strip() if row[0] else ""
                    date_val = None
                    if isinstance(row[1], datetime): date_val = row[1].date()
                    elif isinstance(row[1], date): date_val = row[1]
                    elif isinstance(row[1], str):
                        try: date_val = datetime.strptime(row[1], "%Y-%m-%d").date()
                        except: pass
                        
                    if not date_val: continue
                    
                    emp = None
                    if emp_id_val and emp_id_val.isdigit():
                        emp = session.query(Employee).filter_by(id=int(emp_id_val)).first()
                    if not emp: 
                        errors.append(f"Attendance: Employee ID '{emp_id_val}' not found")
                        continue
                        
                    # Parse Times
                    def parse_t(t_val):
                        if not t_val: return None
                        if isinstance(t_val, time): return datetime.combine(date_val, t_val)
                        if isinstance(t_val, datetime): return t_val # Already datetime?
                        if isinstance(t_val, str):
                            try: return datetime.combine(date_val, datetime.strptime(t_val, "%H:%M").time())
                            except: return None
                        return None
                        
                    cin = parse_t(row[2])
                    cout = parse_t(row[3])
                    
                    # Check duplicate
                    att = session.query(Attendance).filter_by(employee_id=emp.id, date=date_val).first()
                    if not att:
                        att = Attendance(
                            employee_id=emp.id,
                            date=date_val,
                            clock_in=cin,
                            clock_out=cout
                        )
                        session.add(att)
                        count += 1
                session.flush()

            # 10. ShortLeave
            if "ShortLeave" in selected_sheets and "ShortLeave" in wb.sheetnames:
                ws = wb["ShortLeave"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0] and not row[1]: continue
                    # emp_id, date (YYYY-MM-DD), start_time (HH:MM), end_time (HH:MM), reason
                    emp_id_val = str(row[0]).strip() if row[0] else ""
                    date_val = None
                    if isinstance(row[1], datetime): date_val = row[1].date()
                    elif isinstance(row[1], date): date_val = row[1]
                    elif isinstance(row[1], str):
                        try: date_val = datetime.strptime(row[1], "%Y-%m-%d").date()
                        except: pass

                    if not date_val: continue

                    emp = None
                    if emp_id_val and emp_id_val.isdigit():
                        emp = session.query(Employee).filter_by(id=int(emp_id_val)).first()
                    if not emp:
                        errors.append(f"ShortLeave: Employee ID '{emp_id_val}' not found")
                        continue

                    # Parse start/end times
                    def parse_sl_time(t_val):
                        if not t_val: return None
                        if isinstance(t_val, time): return t_val
                        if isinstance(t_val, datetime): return t_val.time()
                        if isinstance(t_val, str):
                            try: return datetime.strptime(t_val, "%H:%M").time()
                            except:
                                try: return datetime.strptime(t_val, "%H:%M:%S").time()
                                except: return None
                        return None

                    start_t = parse_sl_time(row[2])
                    end_t = parse_sl_time(row[3])
                    reason = str(row[4]) if row[4] else ""

                    if not start_t or not end_t:
                        errors.append(f"ShortLeave: Invalid times for {emp_id_val} on {date_val}")
                        continue

                    # Skip duplicates (same employee, date, start_time)
                    exists = session.query(ShortLeave).filter_by(
                        employee_id=emp.id, date=date_val, start_time=start_t
                    ).first()
                    if not exists:
                        sl = ShortLeave(
                            employee_id=emp.id,
                            date=date_val,
                            start_time=start_t,
                            end_time=end_t,
                            reason=reason,
                            status="Pending"
                        )
                        session.add(sl)
                        count += 1
                session.flush()

            # 11. SalaryBreakdown
            if "SalaryBreakdown" in selected_sheets and "SalaryBreakdown" in wb.sheetnames:
                ws = wb["SalaryBreakdown"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0] or not row[1]: continue
                    emp_id_val = str(row[0]).strip()
                    year_val = row[1]
                    valid_to_val = row[2]
                    
                    if isinstance(valid_to_val, datetime): valid_to_val = valid_to_val.date()
                    elif isinstance(valid_to_val, str) and valid_to_val:
                        try: valid_to_val = datetime.strptime(valid_to_val, "%Y-%m-%d").date()
                        except: valid_to_val = None
                    else:
                        valid_to_val = None
                        
                    emp = None
                    if emp_id_val.isdigit():
                        emp = session.query(Employee).filter_by(id=int(emp_id_val)).first()
                        
                    if not emp:
                        errors.append(f"SalaryBreakdown: Employee ID '{emp_id_val}' not found")
                        continue
                        
                    sb = session.query(SalaryBreakdown).filter_by(employee_id=emp.id, year=int(year_val)).first()
                    
                    if not sb:
                        sb = SalaryBreakdown(
                            employee_id=emp.id,
                            year=int(year_val)
                        )
                        session.add(sb)
                        
                    sb.valid_to = valid_to_val
                    sb.basic = float(row[3] or 0.0)
                    sb.house_rent_allowance = float(row[4] or 0.0)
                    sb.conveyance = float(row[5] or 0.0)
                    sb.medical = float(row[6] or 0.0)
                    sb.mobile_bill = float(row[7] or 0.0)
                    sb.transportation_allowance = float(row[8] or 0.0)
                    sb.other_allowance = float(row[9] or 0.0)
                    
                    count += 1
                session.flush()

            session.commit()
            return count, errors
            
        except Exception as e:
            session.rollback()
            return 0, [str(e)]
