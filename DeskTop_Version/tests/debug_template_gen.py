import sys
import os

# Add project root to path
sys.path.append(os.getcwd())

from services.legacy_import_service import LegacyImportService

def test_generate():
    print("Testing generate_template...")
    # Test ALL keys
    table_types = [
        "companies", "business_areas", "designations", "designation_subcategories", 
        "shifts", "employees", "attendance", "weekly_holidays", 
        "leave_quotas", "holiday_calendar", "short_leaves"
    ]
    output_path = "debug_template.xlsx"
    
    try:
        path = LegacyImportService.generate_template(table_types, output_path)
        print(f"Success! Template generated at: {path}")
        
        import openpyxl
        dest_wb = openpyxl.load_workbook(path)
        print(f"Sheet names: {dest_wb.sheetnames}")
        
        if "companies" in dest_wb.sheetnames and "employees" in dest_wb.sheetnames:
            print("Validation Passed: Required sheets exist.")
            
            if "employees" in dest_wb.sheetnames:
                ws = dest_wb["employees"]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    print(f"Row 2 Data: {rows[1]}")
                    # Check for specific content in cell A2 (first column of second row)
                    if str(rows[1][0]) == "500001":
                        print("SUCCESS: Template used data/Import_Template.xlsx (A2 is 500001)")
                    # Check for 'Mehedi' in any cell of the second row
                    elif any('Mehedi' in str(cell) for cell in rows[1]):
                        print("SUCCESS: Template used data/Import_Template.xlsx (Row 2 contains 'Mehedi')")
                    else:
                        print("WARNING: Template used HARDCODED defaults (Fallback active?) - A2 not 500001 and 'Mehedi' not found in row 2.")
                else:
                    print("ERROR: Employees sheet is empty or only has headers.")
        else:
            print("Validation Failed: Missing required sheets (companies or employees).")
            
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_generate()
