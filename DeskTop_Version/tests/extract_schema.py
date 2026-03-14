import openpyxl
import os
import datetime

file_path = os.path.join("data", "Import_Template.xlsx")
wb = openpyxl.load_workbook(file_path, data_only=True)

schemas = {}

def format_val(v):
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M:%S")
    return v

print("TABLE_SCHEMAS = {")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = list(ws.iter_rows(max_row=2, values_only=True))
    if len(rows) >= 1:
        headers = list(rows[0])
        dummy = []
        if len(rows) >= 2:
            dummy = [format_val(c) for c in rows[1]]
        
        print(f'    "{sheet}": {{')
        print(f'        "headers": {headers},')
        print(f'        "dummy": {dummy}')
        print("    },")
print("}")
